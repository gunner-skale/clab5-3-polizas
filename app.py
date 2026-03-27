# app.py - Versión para comparar 2 Excels distintos con IA
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from google import genai
from google.genai import types
from io import BytesIO
import os
import time
from dotenv import load_dotenv
import difflib

# Cargar variables de entorno
load_dotenv()

# ============================================
# CONFIGURACIÓN
# ============================================
st.set_page_config(page_title="Comparador de Pólizas AI - 2 Excels", page_icon="🤖", layout="wide")

# CSS
st.markdown("""
<style>
    .main-header { font-size: 2.5rem; font-weight: bold; color: #1f77b4; text-align: center; margin-bottom: 1rem; }
    .sub-header { font-size: 1.2rem; color: #666; text-align: center; margin-bottom: 2rem; }
    .stButton>button { width: 100%; background-color: #1f77b4; color: white; font-weight: bold; padding: 0.75rem; }
    .stButton>button:hover { background-color: #145a8a; }
    .success-box { background-color: #d4edda; color: #155724; padding: 1rem; border-radius: 0.5rem; border-left: 5px solid #28a745; }
    .error-box { background-color: #f8d7da; color: #721c24; padding: 1rem; border-radius: 0.5rem; border-left: 5px solid #dc3545; }
    .warning-box { background-color: #fff3cd; color: #856404; padding: 1rem; border-radius: 0.5rem; border-left: 5px solid #ffc107; }
    .info-box { background-color: #d1ecf1; color: #0c5460; padding: 1rem; border-radius: 0.5rem; border-left: 5px solid #17a2b8; }
    .mejora-badge { background-color: #C6EFCE; color: #006100; padding: 0.2rem 0.5rem; border-radius: 0.3rem; font-weight: bold; }
    .retroceso-badge { background-color: #FFC7CE; color: #9C0006; padding: 0.2rem 0.5rem; border-radius: 0.3rem; font-weight: bold; }
    .equivalente-badge { background-color: #D9E1F2; color: #0066CC; padding: 0.2rem 0.5rem; border-radius: 0.3rem; font-weight: bold; }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIÓN PARA LOGO EN HEADER
# ============================================


# ============================================
# FUNCIONES UTILITARIAS - SIN CAMBIOS EN IA
# ============================================

def sanitizar_texto(texto):
    """Limpia texto para evitar errores de encoding"""
    if not texto:
        return ""
    texto = str(texto)
    texto = ''.join(char for char in texto if ord(char) >= 32 or char in ['\n', '\t', '\r'])
    texto = ' '.join(texto.split())
    return texto.strip().upper()

def detectar_ok_directo(texto_respuesta):
    """Detecta si la respuesta indica OK sin necesidad de consultar IA"""
    texto = sanitizar_texto(texto_respuesta)

    indicadores_ok = [
        "SE OTORGA", "OTORGA", "SE ACEPTA", "ACEPTA", "CUBRE", "INCLUYE",
        "AMPARA", "PROTEGE", "CONCEDE", "AUTORIZA", "APRUEBA", "SI", "SÍ"
    ]

    indicadores_no = [
        "NO SE OTORGA", "NO OTORGA", "NO SE ACEPTA", "NO ACEPTA",
        "NO CUBRE", "NO INCLUYE", "EXCLUYE", "RECHAZA", "DENIEGA", "NO APLICA"
    ]

    for indicador in indicadores_ok:
        if indicador in texto:
            negado = False
            for neg in indicadores_no:
                if neg in texto:
                    negado = True
                    break
            if not negado:
                return True, f"✅ {indicador.title()}"

    for indicador in indicadores_no:
        if indicador in texto:
            return True, f"DIFERENCIA: {indicador.title()}"

    return False, None

# ============================================
# FUNCIONES DE IA - ⚠️ SIN CAMBIOS (COMO SOLICITASTE)
# ============================================

def inicializar_cliente():
    """Inicializa cliente Gemini desde variable de entorno"""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        st.error("❌ No se encontró GEMINI_API_KEY en el archivo .env")
        st.stop()
    return genai.Client(api_key=api_key)

def comparar_lote_con_mejoras(items_lote, client, max_retries=2):
    """
    Compara lote de items detectando MEJORAS, RETROCESOS y EQUIVALENCIAS
    ⚠️ ESTA FUNCIÓN NO HA SIDO MODIFICADA - LÓGICA DE IA ORIGINAL
    """
    if not items_lote:
        return []

    resultados_preliminares = []
    items_para_ia = []

    # Primero, detección rápida de OK/DIFERENCIA básica
    for idx, orig, resp in items_lote:
        es_ok_directo, motivo = detectar_ok_directo(resp)
        if es_ok_directo:
            if "DIFERENCIA" in motivo:
                resultados_preliminares.append((idx, "DIFERENCIA", motivo, motivo))
            else:
                resultados_preliminares.append((idx, "OK", motivo, "Condiciones equivalentes"))
        else:
            items_para_ia.append((idx, orig, resp))
            resultados_preliminares.append((idx, "PENDIENTE", None, None))

    if not items_para_ia:
        return [(idx, est, obs, detalle) for idx, est, obs, detalle in resultados_preliminares if est != "PENDIENTE"]

    # Prompt mejorado para detectar mejoras y retrocesos
    prompt_items = ""
    mapeo_ia = {}

    for i, (idx, orig, resp) in enumerate(items_para_ia, 1):
        mapeo_ia[i] = idx
        prompt_items += f"""
--- ITEM {i} ---
PÓLIZA ANTERIOR (NUESTRA PROPUESTA): {orig[:500]}
NUEVA PÓLIZA (RESPUESTA ASEGURADORA): {resp[:400]}
"""

    prompt = f"""Actúa como AUDITOR SENIOR DE SEGUROS con 15 años de experiencia. Tu tarea es COMPARAR la póliza anterior con la nueva póliza y determinar si hay MEJORAS, RETROCESOS o son EQUIVALENTES.

{prompt_items}

CRITERIOS DE EVALUACIÓN DETALLADOS:

1. **MEJORA (MEJORA)**: La nueva póliza es SUPERIOR en al menos uno de estos aspectos:
   - Monto asegurado MAYOR
   - Deducible MENOR
   - Cobertura MÁS AMPLIA (incluye más riesgos)
   - Exclusiones REDUCIDAS o ELIMINADAS
   - Condiciones MÁS FAVORABLES para el asegurado
   - Plazo MÁS LARGO
   - Prima MENOR (mismo o mejor cobertura)

2. **RETROCESO (RETROCESO)**: La nueva póliza es INFERIOR en al menos uno de estos aspectos:
   - Monto asegurado MENOR
   - Deducible MAYOR
   - Cobertura MÁS RESTRINGIDA (excluye riesgos que antes cubría)
   - Nuevas exclusiones o condiciones restrictivas
   - Condiciones MENOS FAVORABLES
   - Plazo MÁS CORTO
   - Prima MAYOR (misma o peor cobertura)

3. **EQUIVALENTE (OK)**: Las condiciones son ESENCIALMENTE IGUALES, con diferencias no sustanciales

4. **DIFERENCIA (DIFERENCIA)**: Hay cambios significativos pero no claramente mejor o peor (ej: cambia estructura, no comparable directamente)

REGLAS DE ANÁLISIS:
- Si hay MEJORA en un aspecto pero RETROCESO en otro, prioriza el análisis más relevante para el asegurado
- Para montos: considera significativo >10% de diferencia
- Para exclusiones: cualquier exclusión nueva es un RETROCESO
- Para inclusiones: cualquier cobertura nueva es una MEJORA

FORMATO OBLIGATORIO - UNA LÍNEA POR ITEM CON 3 PARTES SEPARADAS POR "|":
RESULTADO_{i}: TIPO|MENSAJE_CORTO|ANÁLISIS_DETALLADO

Donde:
- TIPO: MEJORA, RETROCESO, OK, DIFERENCIA
- MENSAJE_CORTO: Frase breve de 15-30 palabras resumiendo el cambio
- ANÁLISIS_DETALLADO: Explicación extensa de 50-150 palabras

EJEMPLOS:
RESULTADO_1: MEJORA|Monto asegurado aumentó de $100,000 a $150,000 (+50%)|Se incrementó significativamente la suma asegurada, ofreciendo mayor protección al asegurado sin cambios en las exclusiones.
RESULTADO_2: RETROCESO|Se eliminó cobertura por inundación que sí existía antes|La nueva póliza excluye explícitamente daños por inundación, un riesgo relevante que estaba cubierto en la póliza anterior.
RESULTADO_3: OK|Condiciones equivalentes, solo cambios menores de redacción|La cobertura y condiciones son esencialmente las mismas, no hay cambios sustanciales en montos, deducibles o exclusiones.
RESULTADO_4: DIFERENCIA|Cambio en estructura de coberturas|La nueva póliza reorganiza las coberturas en un formato diferente, difícil de comparar directamente. Se requiere análisis adicional de equivalencia real."""

    for intento in range(max_retries):
        try:
            response = client.models.generate_content(
                model="models/gemini-3.1-flash-lite-preview",
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.1,
                    max_output_tokens=1024,
                    top_p=0.95,
                )
            )

            if not response or not response.text:
                if intento < max_retries - 1:
                    time.sleep(0.5)
                    continue
                for i, (idx, orig, resp) in enumerate(items_para_ia, 1):
                    for j, (id_res, est, obs, det) in enumerate(resultados_preliminares):
                        if id_res == idx and est == "PENDIENTE":
                            resultados_preliminares[j] = (idx, "ERROR", "Sin respuesta API", "No se pudo obtener análisis de la IA")
                break

            lineas = response.text.strip().split('\n')
            resultados_ia = {}

            for linea in lineas:
                linea = linea.strip()
                if not linea.startswith("RESULTADO_"):
                    continue
                try:
                    num_str = linea.split("_")[1].split(":")[0]
                    num = int(num_str)
                    contenido = linea.split(":", 1)[1].strip() if ":" in linea else linea

                    if "|" in contenido:
                        partes = contenido.split("|")
                        tipo = partes[0].strip().upper()
                        mensaje_corto = partes[1].strip() if len(partes) > 1 else "Sin resumen"
                        detalle = partes[2].strip() if len(partes) > 2 else "Sin detalles"
                        if tipo not in ["MEJORA", "RETROCESO", "OK", "DIFERENCIA"]:
                            tipo = "DIFERENCIA"
                        resultados_ia[num] = (tipo, mensaje_corto, detalle)
                    else:
                        resultados_ia[num] = ("DIFERENCIA", contenido[:100], contenido)
                except Exception as e:
                    continue

            for i, (idx, orig, resp) in enumerate(items_para_ia, 1):
                if i in resultados_ia:
                    tipo, mensaje, detalle = resultados_ia[i]
                    for j, (id_res, est, obs, det) in enumerate(resultados_preliminares):
                        if id_res == idx and est == "PENDIENTE":
                            resultados_preliminares[j] = (idx, tipo, mensaje, detalle)
                            break
                else:
                    for j, (id_res, est, obs, det) in enumerate(resultados_preliminares):
                        if id_res == idx and est == "PENDIENTE":
                            resultados_preliminares[j] = (idx, "ERROR", "No procesado", "El item no pudo ser analizado correctamente")
                            break
            break
        except Exception as e:
            if intento < max_retries - 1:
                time.sleep(1)
                continue
            for i, (idx, orig, resp) in enumerate(items_para_ia, 1):
                for j, (id_res, est, obs, det) in enumerate(resultados_preliminares):
                    if id_res == idx and est == "PENDIENTE":
                        resultados_preliminares[j] = (idx, "ERROR", f"Error técnico", f"Error: {str(e)[:100]}")

    return [(idx, tipo, mensaje, detalle) for idx, tipo, mensaje, detalle in resultados_preliminares if tipo != "PENDIENTE"]

# ============================================
# NUEVAS FUNCIONES PARA 2 EXCELS
# ============================================

def encontrar_coincidencias(clausulas_origen, clausulas_destino, umbral_similitud=0.7):
    """
    Encuentra cláusulas similares entre dos listas usando difflib
    Retorna: dict con {index_origen: (index_destino, score_similitud)}
    """
    coincidencias = {}
    usadas_destino = set()

    for i, clausula_orig in enumerate(clausulas_origen):
        mejor_match = None
        mejor_score = 0

        for j, clausula_dest in enumerate(clausulas_destino):
            if j in usadas_destino:
                continue
            # Calcular similitud
            score = difflib.SequenceMatcher(None,
                                          sanitizar_texto(clausula_orig),
                                          sanitizar_texto(clausula_dest)).ratio()
            if score > mejor_score and score >= umbral_similitud:
                mejor_score = score
                mejor_match = j

        if mejor_match is not None:
            coincidencias[i] = (mejor_match, mejor_score)
            usadas_destino.add(mejor_match)

    return coincidencias

def procesar_dos_excels(file_antiguo, file_nuevo, col_antiguo, col_nuevo,
                        fila_inicio, progress_bar, status_text, tamanio_lote=3):
    """
    PROCESA 2 EXCELS DISTINTOS: Compara cláusulas entre póliza antigua y nueva
    """
    # Colores
    VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    AZUL_CLARO = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    AZUL_MEJORA = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    client = inicializar_cliente()

    # Cargar ambos workbooks
    wb_antiguo = load_workbook(file_antiguo)
    wb_nuevo = load_workbook(file_nuevo)

    resultados_totales = []
    hojas_procesadas = []

    # Obtener hojas que empiecen con "PP" en ambos archivos
    hojas_antiguo = [name for name in wb_antiguo.sheetnames if name.startswith("PP")]
    hojas_nuevo = [name for name in wb_nuevo.sheetnames if name.startswith("PP")]

    # Procesar solo hojas que existan en ambos archivos
    hojas_comunes = set(hojas_antiguo) & set(hojas_nuevo)

    if not hojas_comunes:
        status_text.text("❌ No hay hojas 'PP' comunes en ambos archivos")
        return wb_antiguo, [], "No hay hojas comunes para comparar"

    total_hojas = len(hojas_comunes)

    for idx_hoja, sheet_name in enumerate(sorted(hojas_comunes), 1):
        ws_antiguo = wb_antiguo[sheet_name]
        ws_nuevo = wb_nuevo[sheet_name]

        # Crear columnas de resultado en el Excel ANTIGUO (base)
        col_resultado = ws_antiguo.max_column + 1
        col_tipo_cambio = ws_antiguo.max_column + 2
        col_coincidencia = ws_antiguo.max_column + 3
        col_observacion_detallada = ws_antiguo.max_column + 4

        # Headers en el Excel base
        headers = [
            ("🔍 RESULTADO COMPARACIÓN", col_resultado),
            ("📊 TIPO DE CAMBIO", col_tipo_cambio),
            ("🔗 CLÁUSULA SIMILAR EN NUEVA", col_coincidencia),
            ("📝 ANÁLISIS DETALLADO IA", col_observacion_detallada)
        ]

        for header_value, col_num in headers:
            header_cell = ws_antiguo.cell(row=fila_inicio-1, column=col_num)
            header_cell.value = header_value
            header_cell.font = Font(bold=True, size=10, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            header_cell.alignment = Alignment(wrap_text=True)

        # Recolectar datos de ambos Excels
        datos_antiguo = []
        for row in range(fila_inicio, ws_antiguo.max_row + 1):
            val = str(ws_antiguo.cell(row=row, column=col_antiguo).value or "").strip()
            if val and len(val) > 5:
                datos_antiguo.append((row, val))

        datos_nuevo = []
        for row in range(fila_inicio, ws_nuevo.max_row + 1):
            val = str(ws_nuevo.cell(row=row, column=col_nuevo).value or "").strip()
            if val and len(val) > 5:
                datos_nuevo.append((row, val))

        if not datos_antiguo:
            continue

        # Encontrar coincidencias entre cláusulas
        clausulas_antiguo = [d[1] for d in datos_antiguo]
        clausulas_nuevo = [d[1] for d in datos_nuevo]

        coincidencias = encontrar_coincidencias(clausulas_antiguo, clausulas_nuevo)

        # Preparar items para IA: (row_antiguo, texto_antiguo, texto_nuevo_o_vacio)
        items_para_comparar = []
        filas_sin_coincidencia = []

        for idx_antiguo, (row_ant, texto_ant) in enumerate(datos_antiguo):
            if idx_antiguo in coincidencias:
                idx_nuevo, score = coincidencias[idx_antiguo]
                texto_nuevo = datos_nuevo[idx_nuevo][1]
                items_para_comparar.append((row_ant, texto_ant, texto_nuevo))
            else:
                filas_sin_coincidencia.append((row_ant, texto_ant, "⚠️ NO ENCONTRADA EN NUEVA PÓLIZA"))

        # Agregar cláusulas nuevas que no existen en la antigua (al final)
        items_nuevos_sin_igual = []
        for idx_nuevo, (row_nue, texto_nue) in enumerate(datos_nuevo):
            encontrado = False
            for idx_ant, (match_idx, score) in coincidencias.items():
                if match_idx == idx_nuevo:
                    encontrado = True
                    break
            if not encontrado:
                items_nuevos_sin_igual.append((texto_nue, "🆕 NUEVA - NO EXISTÍA EN ANTIGUA"))

        total_items = len(items_para_comparar) + len(filas_sin_coincidencia) + len(items_nuevos_sin_igual)
        procesados = 0

        # Procesar items con coincidencia usando IA
        if items_para_comparar:
            for i in range(0, len(items_para_comparar), tamanio_lote):
                lote = items_para_comparar[i:i + tamanio_lote]
                filas_str = ", ".join([str(r) for r, _, _ in lote])
                status_text.text(f"Hoja {idx_hoja}/{total_hojas} '{sheet_name}': Comparando filas {filas_str}")

                resultados_lote = comparar_lote_con_mejoras(lote, client)

                for resultado_analisis in resultados_lote:
                    idx, tipo, mensaje_corto, detalle_largo = resultado_analisis

                    # Aplicar formato en Excel ANTIGUO
                    cell_res = ws_antiguo.cell(row=idx, column=col_resultado)
                    cell_tipo = ws_antiguo.cell(row=idx, column=col_tipo_cambio)
                    cell_coinc = ws_antiguo.cell(row=idx, column=col_coincidencia)
                    cell_detalle = ws_antiguo.cell(row=idx, column=col_observacion_detallada)

                    # Encontrar texto similar en nueva póliza para mostrar
                    texto_coincidencia = ""
                    for row_ant, txt_ant, txt_nue in items_para_comparar:
                        if row_ant == idx:
                            texto_coincidencia = txt_nue[:200] + "..." if len(txt_nue) > 200 else txt_nue
                            break

                    if tipo == "MEJORA":
                        cell_res.value = f"✅ MEJORA: {mensaje_corto}"
                        cell_res.fill = AZUL_MEJORA
                        cell_res.font = Font(color="003366", bold=True, size=9)
                        cell_tipo.value = "📈 MEJORA SIGNIFICATIVA"
                        cell_tipo.fill = AZUL_MEJORA
                        cell_tipo.font = Font(color="003366", bold=True)
                    elif tipo == "RETROCESO":
                        cell_res.value = f"⚠️ RETROCESO: {mensaje_corto}"
                        cell_res.fill = ROJO
                        cell_res.font = Font(color="9C0006", bold=True, size=9)
                        cell_tipo.value = "📉 RETROCESO EN CONDICIONES"
                        cell_tipo.fill = ROJO
                        cell_tipo.font = Font(color="9C0006", bold=True)
                    elif tipo == "OK":
                        cell_res.value = f"✅ EQUIVALENTE: {mensaje_corto}"
                        cell_res.fill = VERDE
                        cell_res.font = Font(color="006100", bold=True, size=9)
                        cell_tipo.value = "🔄 CONDICIONES EQUIVALENTES"
                        cell_tipo.fill = VERDE
                        cell_tipo.font = Font(color="006100", bold=True)
                    elif tipo == "DIFERENCIA":
                        cell_res.value = f"⚠️ DIFERENCIA: {mensaje_corto}"
                        cell_res.fill = AMARILLO
                        cell_res.font = Font(color="9C5700", bold=True, size=9)
                        cell_tipo.value = "📊 DIFERENCIA SIGNIFICATIVA"
                        cell_tipo.fill = AMARILLO
                        cell_tipo.font = Font(color="9C5700", bold=True)
                    else:
                        cell_res.value = f"❌ ERROR: {mensaje_corto}"
                        cell_res.fill = AMARILLO
                        cell_res.font = Font(color="9C5700", bold=True, size=9)
                        cell_tipo.value = "⚠️ ERROR PROCESAMIENTO"
                        cell_tipo.fill = AMARILLO
                        cell_tipo.font = Font(color="9C5700", bold=True)

                    cell_coinc.value = texto_coincidencia
                    cell_coinc.alignment = Alignment(wrap_text=True, vertical="top")
                    cell_coinc.font = Font(size=8, italic=True)

                    cell_detalle.value = detalle_largo
                    cell_detalle.alignment = Alignment(wrap_text=True, vertical="top")
                    cell_detalle.font = Font(size=8)

                    resultados_totales.append({
                        'hoja': sheet_name,
                        'fila_antigua': idx,
                        'tipo': tipo,
                        'resumen': mensaje_corto,
                        'detalle': detalle_largo,
                        'coincidencia': 'Sí'
                    })

                procesados += len(lote)
                progress_bar.progress(min((idx_hoja - 1 + procesados/total_items) / total_hojas, 0.95))
                time.sleep(0.1)

        # Procesar filas sin coincidencia (cláusulas que desaparecieron)
        for row_ant, texto_ant, mensaje in filas_sin_coincidencia:
            cell_res = ws_antiguo.cell(row=row_ant, column=col_resultado)
            cell_tipo = ws_antiguo.cell(row=row_ant, column=col_tipo_cambio)
            cell_coinc = ws_antiguo.cell(row=row_ant, column=col_coincidencia)
            cell_detalle = ws_antiguo.cell(row=row_ant, column=col_observacion_detallada)

            cell_res.value = f"❌ ELIMINADA: {mensaje}"
            cell_res.fill = ROJO
            cell_res.font = Font(color="9C0006", bold=True, size=9)

            cell_tipo.value = "🗑️ CLÁUSULA ELIMINADA"
            cell_tipo.fill = ROJO
            cell_tipo.font = Font(color="9C0006", bold=True)

            cell_coinc.value = "No encontrada en nueva póliza"
            cell_coinc.font = Font(size=8, italic=True, color="999999")

            cell_detalle.value = f"Esta cláusula existía en la póliza anterior pero NO aparece en la nueva póliza. Verificar si fue intencional o requiere negociación."
            cell_detalle.alignment = Alignment(wrap_text=True, vertical="top")
            cell_detalle.font = Font(size=8)

            resultados_totales.append({
                'hoja': sheet_name,
                'fila_antigua': row_ant,
                'tipo': 'ELIMINADA',
                'resumen': mensaje,
                'detalle': 'Cláusula eliminada en nueva póliza',
                'coincidencia': 'No'
            })
            procesados += 1

        # Agregar cláusulas nuevas al FINAL del Excel antiguo
        if items_nuevos_sin_igual:
            fila_inicio_nuevas = ws_antiguo.max_row + 2

            # Header para sección de nuevas cláusulas
            header_nuevas = ws_antiguo.cell(row=fila_inicio_nuevas, column=1)
            header_nuevas.value = "🆕 CLÁUSULAS NUEVAS (No existían en póliza anterior)"
            header_nuevas.font = Font(bold=True, size=11, color="FFFFFF")
            header_nuevas.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws_antiguo.merge_cells(start_row=fila_inicio_nuevas, end_row=fila_inicio_nuevas,
                                  start_column=1, end_column=col_observacion_detallada)

            for idx_nueva, (texto_nuevo, mensaje) in enumerate(items_nuevos_sin_igual, start=fila_inicio_nuevas+1):
                # Preparar item para IA comparando con vacío
                items_ia = [(idx_nueva, "NO EXISTÍA ANTERIORMENTE", texto_nuevo)]
                resultados_ia = comparar_lote_con_mejoras(items_ia, client)

                if resultados_ia:
                    _, tipo, mensaje_corto, detalle_largo = resultados_ia[0]
                else:
                    tipo, mensaje_corto, detalle_largo = "NUEVA", "Cláusula nueva agregada", texto_nuevo[:200]

                ws_antiguo.cell(row=idx_nueva, column=1).value = f"🆕 {texto_nuevo[:150]}..."
                ws_antiguo.cell(row=idx_nueva, column=1).font = Font(italic=True, size=9)

                cell_res = ws_antiguo.cell(row=idx_nueva, column=col_resultado)
                cell_res.value = f"✨ {mensaje}"
                cell_res.fill = AZUL_CLARO
                cell_res.font = Font(color="0066CC", bold=True, size=9)

                cell_tipo = ws_antiguo.cell(row=idx_nueva, column=col_tipo_cambio)
                cell_tipo.value = "🆕 CLÁUSULA NUEVA"
                cell_tipo.fill = AZUL_CLARO
                cell_tipo.font = Font(color="0066CC", bold=True)

                cell_coinc = ws_antiguo.cell(row=idx_nueva, column=col_coincidencia)
                cell_coinc.value = "N/A - Nueva cláusula"
                cell_coinc.font = Font(size=8, italic=True)

                cell_detalle = ws_antiguo.cell(row=idx_nueva, column=col_observacion_detallada)
                cell_detalle.value = f"{detalle_largo}\n\n💡 Recomendación: Evaluar si esta nueva cláusula beneficia al cliente o requiere aclaración."
                cell_detalle.alignment = Alignment(wrap_text=True, vertical="top")
                cell_detalle.font = Font(size=8)

                resultados_totales.append({
                    'hoja': sheet_name,
                    'fila_antigua': f'Nueva-{idx_nueva}',
                    'tipo': 'NUEVA',
                    'resumen': mensaje_corto,
                    'detalle': detalle_largo,
                    'coincidencia': 'Nueva'
                })
                procesados += 1

        # Ajustar anchos de columna
        from openpyxl.utils import get_column_letter
        ws_antiguo.column_dimensions[get_column_letter(col_resultado)].width = 40
        ws_antiguo.column_dimensions[get_column_letter(col_tipo_cambio)].width = 28
        ws_antiguo.column_dimensions[get_column_letter(col_coincidencia)].width = 50
        ws_antiguo.column_dimensions[get_column_letter(col_observacion_detallada)].width = 70

        hojas_procesadas.append(f"{sheet_name} ({len(datos_antiguo)} cláusulas)")
        progress_bar.progress(min(idx_hoja / total_hojas, 0.99))

    progress_bar.progress(1.0)
    status_text.text(f"✅ {len(hojas_procesadas)} hojas comparadas | {len(resultados_totales)} cláusulas analizadas")

    return wb_antiguo, resultados_totales, "Procesamiento completado"

def agregar_logo_header():
    """
    Inserta el logo en la esquina superior derecha usando HTML/CSS
    El logo debe estar en la misma carpeta del proyecto
    """
    st.markdown("""
    <style>
        .logo-container {
            position: fixed;
            top: 15px;
            right: 20px;
            z-index: 9999;
            display: flex;
            align-items: center;
        }
        .logo-img {
            max-height: 50px;
            max-width: 150px;
            object-fit: contain;
            border-radius: 4px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        /* Ajustar padding del header para que no tape contenido */
        .main .block-container {
            padding-top: 2rem;
        }
    </style>
    <div class="logo-container">
        <img src="./logoCL1.jpeg" alt="Logo" class="logo-img">
    </div>
    """, unsafe_allow_html=True)

# ============================================
# INTERFAZ PRINCIPAL
# ============================================

def main():
    st.markdown('<div class="main-header">🤖 Comparador Inteligente de Pólizas Antiguas y Nuevas</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Compara póliza ANTIGUA vs NUEVA identificando mejoras, retrocesos y diferencias</div>', unsafe_allow_html=True)

    # Verificar API Key
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        st.markdown('<div class="error-box"><strong>❌ Error:</strong> Crear archivo <code>.env</code> con GEMINI_API_KEY=tu_key</div>', unsafe_allow_html=True)
        st.stop()

    st.markdown('<div class="success-box"><strong>✅ API Key configurada</strong> | Gemini AI activa para análisis</div>', unsafe_allow_html=True)

    # Sidebar - Configuración
    st.sidebar.header("⚙️ Configuración de Comparación")

    st.sidebar.info("""
    **📋 Instrucciones:**
    1. Sube la póliza ANTIGUA (archivo base)
    2. Sube la póliza NUEVA (a comparar)
    3. Selecciona las columnas con las cláusulas
    4. ¡Compara y obtén análisis con IA!
    """)

    # Selector de columnas para cada Excel
    col_options = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"]

    st.sidebar.subheader("📄 PÓLIZA ANTIGUA (Base)")
    col_antiguo = st.sidebar.selectbox("Columna con cláusulas", col_options, index=2, key="col_ant")
    col_antiguo_num = ord(col_antiguo) - ord('A') + 1

    st.sidebar.subheader("📄 PÓLIZA NUEVA (Comparar)")
    col_nuevo = st.sidebar.selectbox("Columna con cláusulas", col_options, index=4, key="col_nue")
    col_nuevo_num = ord(col_nuevo) - ord('A') + 1

    fila_inicio = st.sidebar.number_input("🔢 Fila donde inician los datos", min_value=1, value=16, help="Número de fila donde comienza el listado de cláusulas")

    tamanio_lote = st.sidebar.selectbox("📦 Tamaño de lote para IA",
                                       ["Individual (1)", "Parejas (2)", "Tríos (3)", "Cuartetos (4)"],
                                       index=2)
    tamanio_lote_num = int(tamanio_lote.split("(")[1].split(")")[0])

    umbral_similitud = st.sidebar.slider("🎯 Sensibilidad de coincidencia",
                                        min_value=0.5, max_value=0.95, value=0.7, step=0.05,
                                        help="Valor más alto = coincidencias más estrictas")

    st.sidebar.markdown("---")
    st.sidebar.markdown("""
    **🎨 Leyenda de Resultados:**
    - 🟦 <span class="mejora-badge">MEJORA</span>: Nueva póliza es superior
    - 🟥 <span class="retroceso-badge">RETROCESO</span>: Nueva póliza es inferior
    - 🟩 <span class="equivalente-badge">EQUIVALENTE</span>: Condiciones similares
    - 🟨 DIFERENCIA: Cambio no clasificable
    - ❌ ELIMINADA: No existe en nueva póliza
    - ✨ NUEVA: No existía en póliza antigua
    """, unsafe_allow_html=True)

    # Sección de carga de archivos
    st.subheader("📁 Subir Archivos Excel")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 📄 Póliza ANTIGUA")
        uploaded_antiguo = st.file_uploader("Excel con póliza anterior (.xlsx)",
                                           type=['xlsx', 'xls'],
                                           key="antiguo",
                                           help="Este será el archivo base donde se agregarán los resultados")
        if uploaded_antiguo:
            try:
                df_prev = pd.read_excel(uploaded_antiguo, header=None).head(5)
                with st.expander("👁️ Vista previa"):
                    st.dataframe(df_prev, use_container_width=True)
                uploaded_antiguo.seek(0)
            except:
                st.error("Error leyendo archivo")

    with col2:
        st.markdown("### 📄 Póliza NUEVA")
        uploaded_nuevo = st.file_uploader("Excel con nueva póliza (.xlsx)",
                                         type=['xlsx', 'xls'],
                                         key="nuevo",
                                         help="Este archivo se usará para comparar cláusulas")
        if uploaded_nuevo:
            try:
                df_prev = pd.read_excel(uploaded_nuevo, header=None).head(5)
                with st.expander("👁️ Vista previa"):
                    st.dataframe(df_prev, use_container_width=True)
                uploaded_nuevo.seek(0)
            except:
                st.error("Error leyendo archivo")

    # Botón de comparación
    if uploaded_antiguo and uploaded_nuevo:
        if st.button("🚀 INICIAR COMPARACIÓN CON IA", type="primary", use_container_width=True):

            progress_bar = st.progress(0)
            status_text = st.empty()
            results_container = st.empty()

            with st.spinner("🔍 Analizando cláusulas con Inteligencia Artificial..."):
                try:
                    wb_resultado, resultados, mensaje = procesar_dos_excels(
                        uploaded_antiguo, uploaded_nuevo,
                        col_antiguo_num, col_nuevo_num,
                        fila_inicio, progress_bar, status_text,
                        tamanio_lote=tamanio_lote_num
                    )

                    if not resultados:
                        st.warning("⚠️ No se encontraron cláusulas para comparar. Verifica las columnas y filas seleccionadas.")
                        st.stop()

                    # Guardar resultado
                    output = BytesIO()
                    wb_resultado.save(output)
                    output.seek(0)

                    # Estadísticas
                    stats = {
                        'MEJORA': sum(1 for r in resultados if r['tipo'] == 'MEJORA'),
                        'RETROCESO': sum(1 for r in resultados if r['tipo'] == 'RETROCESO'),
                        'OK': sum(1 for r in resultados if r['tipo'] == 'OK'),
                        'DIFERENCIA': sum(1 for r in resultados if r['tipo'] == 'DIFERENCIA'),
                        'ELIMINADA': sum(1 for r in resultados if r['tipo'] == 'ELIMINADA'),
                        'NUEVA': sum(1 for r in resultados if r['tipo'] == 'NUEVA'),
                        'ERROR': sum(1 for r in resultados if r['tipo'] == 'ERROR')
                    }

                    # Mostrar resumen ejecutivo
                    st.success(f"✅ Comparación completada: {len(resultados)} cláusulas analizadas")

                    # Métricas en columnas
                    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                    with col_m1:
                        st.metric("📈 Mejoras", stats['MEJORA'],
                               delta="✅ Favorable" if stats['MEJORA'] > 0 else None,
                               delta_color="normal")
                    with col_m2:
                        st.metric("📉 Retrocesos", stats['RETROCESO'],
                               delta="⚠️ Revisar" if stats['RETROCESO'] > 0 else None,
                               delta_color="inverse")
                    with col_m3:
                        st.metric("🔄 Equivalentes", stats['OK'])
                    with col_m4:
                        st.metric("❌ Eliminadas", stats['ELIMINADA'],
                               delta="⚠️ Verificar" if stats['ELIMINADA'] > 0 else None,
                               delta_color="inverse")

                    # Recomendación para el cliente
                    with st.expander("💡 RECOMENDACIÓN PARA EL CLIENTE", expanded=True):
                        if stats['MEJORA'] > stats['RETROCESO'] and stats['ELIMINADA'] == 0:
                            st.markdown("""
                            <div class="success-box">
                            <strong>🎯 Recomendación: ACEPTAR la nueva póliza</strong><br>
                            La nueva propuesta presenta más mejoras que retrocesos y no elimina coberturas importantes.
                            </div>
                            """, unsafe_allow_html=True)
                        elif stats['RETROCESO'] > stats['MEJORA'] or stats['ELIMINADA'] > 0:
                            st.markdown("""
                            <div class="warning-box">
                            <strong>⚠️ Recomendación: NEGOCIAR antes de aceptar</strong><br>
                            Se identificaron retrocesos o cláusulas eliminadas. Sugiere revisar estos puntos con la aseguradora.
                            </div>
                            """, unsafe_allow_html=True)
                        else:
                            st.markdown("""
                            <div class="info-box">
                            <strong>📊 Recomendación: EVALUAR caso por caso</strong><br>
                            Las condiciones son mayormente equivalentes. Revisar las diferencias específicas según necesidades del cliente.
                            </div>
                            """, unsafe_allow_html=True)

                    # Tabla de resultados detallada
                    with st.expander("📋 Detalle completo de comparación"):
                        df_resultados = pd.DataFrame(resultados)
                        # Filtros
                        filtro_tipo = st.multiselect("Filtrar por tipo:",
                                                   options=df_resultados['tipo'].unique(),
                                                   default=df_resultados['tipo'].unique())
                        df_filtrado = df_resultados[df_resultados['tipo'].isin(filtro_tipo)]
                        st.dataframe(df_filtrado, use_container_width=True)

                    # Botón de descarga
                    nombre_archivo = f"COMPARACION_{uploaded_antiguo.name}".replace('.xlsx', '_vs_NUEVA.xlsx')
                    st.download_button(
                        label="📥 DESCARGAR EXCEL CON ANÁLISIS COMPLETO",
                        data=output,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

                    st.balloons()

                except Exception as e:
                    st.error(f"❌ Error en procesamiento: {str(e)}")
                    st.exception(e)
                    import traceback
                    st.code(traceback.format_exc())

    else:
        st.info("👆 Sube ambos archivos Excel para comenzar la comparación")

if __name__ == "__main__":
    agregar_logo_header()
    main()
